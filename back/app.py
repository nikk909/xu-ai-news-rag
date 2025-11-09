from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask_mail import Mail
from config import Config
from models import db, User, VerificationCode
from utils import (
    create_verification_code,
    send_verification_email,
    send_notification_email,
    verify_code,
    generate_token,
    verify_token
)
from knowledge_base import KnowledgeBase
from ollama_client import OllamaClient
from news_crawler import NewsCrawler
from scheduler import NewsScheduler
from web_search import WebSearcher
from file_processor import FileProcessor
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime
from werkzeug.utils import secure_filename
from pathlib import Path
import os
import logging
import re
from collections import Counter
import jieba
import threading
import time
import uuid
from datetime import datetime, timedelta

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config.from_object(Config)

# 初始化扩展
CORS(app)
db.init_app(app)
mail = Mail(app)

# 初始化知识库和模型
kb = KnowledgeBase()
ollama_client = OllamaClient(model_name='qwen2.5:4b')
web_searcher = WebSearcher()
file_processor = FileProcessor(upload_dir='uploads')

# 创建数据库表
with app.app_context():
    db.create_all()

# 启动定时任务
scheduler = NewsScheduler(app, mail, kb, ollama_client)
scheduler.start()

# 分析任务存储（内存中，实际应用可用Redis）
analysis_tasks = {}
analysis_tasks_lock = threading.Lock()


@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查"""
    return jsonify({'status': 'ok', 'message': '后端服务运行正常'})


@app.route('/api/auth/send-code', methods=['POST'])
def send_code():
    """发送验证码"""
    try:
        data = request.get_json()
        email = data.get('email')
        
        if not email:
            return jsonify({'error': '邮箱地址不能为空'}), 400
        
        # 检查邮箱是否已注册
        existing_user = User.query.filter_by(email=email).first()
        if existing_user:
            return jsonify({'error': '该邮箱已注册'}), 400
        
        # 生成验证码
        code = create_verification_code(email)
        
        # 发送邮件
        if app.config['MAIL_USERNAME']:
            success = send_verification_email(mail, email, code)
            if not success:
                error_msg = '发送邮件失败，请检查邮箱配置'
                # 检查配置
                if not app.config['MAIL_PASSWORD']:
                    error_msg += '（MAIL_PASSWORD未配置）'
                print(f"邮件配置检查: MAIL_SERVER={app.config['MAIL_SERVER']}, MAIL_USERNAME={app.config['MAIL_USERNAME']}, MAIL_PASSWORD={'已配置' if app.config['MAIL_PASSWORD'] else '未配置'}")
                return jsonify({'error': error_msg}), 500
            
            return jsonify({
                'message': '验证码已发送到您的邮箱',
                'email': email
            })
        else:
            # 开发模式：直接返回验证码（生产环境应删除）
            return jsonify({
                'message': '验证码已生成（开发模式）',
                'code': code,  # 仅用于开发测试
                'email': email,
                'warning': '邮件服务未配置，这是开发模式的验证码'
            })
    
    except Exception as e:
        return jsonify({'error': f'服务器错误: {str(e)}'}), 500


@app.route('/api/auth/register', methods=['POST'])
def register():
    """用户注册"""
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        code = data.get('code')
        
        if not email or not password or not code:
            return jsonify({'error': '邮箱、密码和验证码不能为空'}), 400
        
        # 验证码校验
        if not verify_code(email, code):
            return jsonify({'error': '验证码无效或已过期'}), 400
        
        # 检查邮箱是否已注册
        existing_user = User.query.filter_by(email=email).first()
        if existing_user:
            return jsonify({'error': '该邮箱已注册'}), 400
        
        # 创建用户
        user = User(email=email, is_verified=True)
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        # 生成 token
        token = generate_token(user.id, user.email)
        
        return jsonify({
            'message': '注册成功',
            'token': token,
            'user': user.to_dict()
        }), 201
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'服务器错误: {str(e)}'}), 500


@app.route('/api/auth/login', methods=['POST'])
def login():
    """用户登录"""
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        
        if not email or not password:
            return jsonify({'error': '邮箱和密码不能为空'}), 400
        
        # 查找用户
        user = User.query.filter_by(email=email).first()
        
        if not user or not user.check_password(password):
            return jsonify({'error': '邮箱或密码错误'}), 401
        
        if not user.is_verified:
            return jsonify({'error': '账号未验证，请先完成注册验证'}), 403
        
        # 生成 token
        token = generate_token(user.id, user.email)
        
        return jsonify({
            'message': '登录成功',
            'token': token,
            'user': user.to_dict()
        })
    
    except Exception as e:
        return jsonify({'error': f'服务器错误: {str(e)}'}), 500


@app.route('/api/auth/verify-token', methods=['POST'])
def verify_user_token():
    """验证 token"""
    try:
        data = request.get_json()
        token = data.get('token')
        
        if not token:
            return jsonify({'error': 'Token 不能为空'}), 400
        
        payload = verify_token(token)
        if not payload:
            return jsonify({'error': 'Token 无效或已过期'}), 401
        
        # 查找用户
        user = User.query.get(payload['user_id'])
        if not user:
            return jsonify({'error': '用户不存在'}), 404
        
        return jsonify({
            'valid': True,
            'user': user.to_dict()
        })
    
    except Exception as e:
        return jsonify({'error': f'服务器错误: {str(e)}'}), 500


@app.route('/api/knowledge/search', methods=['POST'])
def search_knowledge():
    """搜索知识库，如果无结果则联网搜索"""
    try:
        data = request.get_json()
        query = data.get('query', '')
        page = int(data.get('page', 1))
        page_size = int(data.get('page_size', 10))
        selected_kbs = data.get('selected_kbs', [])  # 用户选择的知识库列表
        
        if not query:
            return jsonify({'error': '搜索关键词不能为空'}), 400
        
        # 获取所有知识库列表
        uploads_dir = file_processor.upload_dir
        all_kbs = []
        if uploads_dir.exists():
            for item in uploads_dir.iterdir():
                if item.is_dir():
                    all_kbs.append(item.name)
        
        # 确定要搜索的知识库
        kbs_to_search = []
        if selected_kbs and len(selected_kbs) > 0:
            # 使用用户选择的知识库（最多5个）
            kbs_to_search = selected_kbs[:5]
        else:
            # 如果没有选择知识库，搜索默认知识库（包含定时任务采集的新闻）
            kbs_to_search = ['default']
        
        # 从多个知识库搜索并合并结果
        kb_results = []
        kb_results_by_source = {}  # 记录每个知识库的结果数量
        
        if kbs_to_search:
            # 搜索选定的知识库
            for kb_name in kbs_to_search:
                try:
                    # default知识库使用不同的路径
                    if kb_name == 'default':
                        kb_instance = KnowledgeBase(db_path='instance/faiss_index_default')
                    else:
                        kb_instance = KnowledgeBase(db_path=f'instance/faiss_index_{kb_name}')
                    
                    # 检查知识库是否有数据
                    kb_stats = kb_instance.get_stats()
                    logger.info(f"知识库 {kb_name} 统计: {kb_stats}")
                    
                    if kb_stats['total_documents'] == 0:
                        logger.warning(f"知识库 {kb_name} 为空，没有可搜索的内容")
                        kb_results_by_source[kb_name] = 0
                        del kb_instance
                        continue
                    
                    # 进一步降低相似度阈值，让更多结果能够返回（从0.1降到0.05）
                    # 同时增加搜索数量，确保能找到相关内容
                    results = kb_instance.search(query, top_k=page_size * 5, similarity_threshold=0.05)
                    logger.info(f"知识库 {kb_name} 搜索 '{query}' 找到 {len(results)} 条结果")
                    
                    # 如果仍然没有结果，尝试更宽松的搜索（不设阈值）
                    if len(results) == 0:
                        logger.info(f"知识库 {kb_name} 使用更宽松的搜索策略")
                        results = kb_instance.search(query, top_k=page_size * 5, similarity_threshold=0.0)
                        logger.info(f"知识库 {kb_name} 宽松搜索找到 {len(results)} 条结果")
                    
                    # 为结果添加知识库来源标记
                    for result in results:
                        result['kb_name'] = kb_name
                        result['from_user_kb'] = (kb_name != 'default')
                        # 确保有来源信息：如果没有source，使用知识库名称
                        if not result.get('metadata'):
                            result['metadata'] = {}
                        if not result['metadata'].get('source'):
                            result['metadata']['source'] = f'知识库: {kb_name}'
                        # 如果没有title，使用知识库名称
                        if not result['metadata'].get('title'):
                            result['metadata']['title'] = f'来自知识库 {kb_name}'
                    
                    kb_results.extend(results)
                    kb_results_by_source[kb_name] = len(results)
                    del kb_instance
                except Exception as e:
                    logger.error(f"搜索知识库 {kb_name} 失败: {e}", exc_info=True)
                    kb_results_by_source[kb_name] = 0
            
            # 按相似度排序（优先显示用户创建的知识库的结果）
            kb_results.sort(key=lambda x: (
                not x.get('from_user_kb', False),  # 用户创建的知识库优先
                -x.get('similarity', 0)  # 然后按相似度降序
            ))
        
        # 如果知识库没有结果，触发联网搜索
        web_results = []
        used_web_search = False
        
        # 记录搜索统计信息
        total_kb_docs = sum(kb_results_by_source.values())
        logger.info(f"知识库搜索完成: 查询='{query}', 选择的知识库={kbs_to_search}, 找到结果数={total_kb_docs}")
        
        # 如果知识库没有结果，触发联网搜索
        if len(kb_results) == 0:
            logger.info(f"知识库无结果，触发联网搜索: {query}")
            used_web_search = True
            
            # 联网搜索（使用真实搜索API，不返回虚拟内容）
            web_search_results = web_searcher.search(query, max_results=3)
            
            if web_search_results:
                # 使用 Ollama 对联网搜索结果进行推理和总结
                for web_result in web_search_results:
                    try:
                        # 使用 Ollama 生成摘要和推理
                        context = f"标题：{web_result['title']}\n内容：{web_result['content']}"
                        summary = ollama_client.answer_question(
                            question=query,
                            context=context
                        )
                        
                        # 构建结果
                        web_results.append({
                            'text': summary,
                            'metadata': {
                                'title': web_result['title'],
                                'source': web_result['source'],
                                'link': web_result.get('link', ''),
                                'published': datetime.now().isoformat(),
                                'original_content': web_result['content'],
                            },
                            'similarity': 0.8,  # 联网搜索结果默认相似度
                            'rank': web_result.get('rank', 1),
                            'from_web': True  # 标记来自联网搜索
                        })
                    except Exception as e:
                        logger.error(f"Ollama处理搜索结果失败: {e}")
                        # 如果 Ollama 处理失败，直接使用原始结果
                        web_results.append({
                            'text': web_result['content'],
                            'metadata': {
                                'title': web_result['title'],
                                'source': web_result['source'],
                                'link': web_result.get('link', ''),
                                'published': datetime.now().isoformat(),
                            },
                            'similarity': 0.7,
                            'rank': web_result.get('rank', 1),
                            'from_web': True
                        })
            else:
                logger.warning(f"联网搜索未找到真实结果，不返回虚拟内容")
        
        # 合并结果：优先显示知识库结果，然后是联网搜索结果
        all_results = kb_results + web_results
        total = len(all_results)
        
        # 添加搜索统计信息到响应中
        search_info = {
            'kb_results_count': len(kb_results),
            'web_results_count': len(web_results),
            'kb_results_by_source': kb_results_by_source,
            'used_web_search': used_web_search
        }
        
        # 分页
        start = (page - 1) * page_size
        end = start + page_size
        paginated_results = all_results[start:end]
        
        # 尝试发送邮件通知给当前登录用户（只有用户明确选择时才发送）
        email_sent = False
        user_email = None
        send_email = data.get('send_email', False)  # 默认不发送，只有用户明确选择时才发送
        
        # 过滤掉admin默认邮箱，避免发送到无效邮箱
        ADMIN_EMAILS = ['admin@xu-news.com', 'admin@example.com']
        
        if total > 0 and send_email:
            try:
                # 从请求头获取token
                auth_header = request.headers.get('Authorization', '')
                if auth_header.startswith('Bearer '):
                    token = auth_header.split(' ')[1]
                    payload = verify_token(token)
                    if payload:
                        user = User.query.get(payload['user_id'])
                        if user and user.is_verified:
                            user_email = user.email
                            
                            # 如果没有从token获取到，尝试从请求体获取
                            if not user_email:
                                user_email = data.get('user_email')
                            
                            # 确保不是admin邮箱，且邮箱有效
                            if user_email and user_email not in ADMIN_EMAILS and '@' in user_email:
                                try:
                                    source_info = "知识库" if not used_web_search else "联网搜索"
                                    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                    subject = f"搜索结果通知 - {query[:30]}{'...' if len(query) > 30 else ''} ({total}条结果)"
                                    
                                    # 构建结果摘要表格
                                    results_summary = ""
                                    for idx, result in enumerate(paginated_results[:10], 1):  # 最多显示10条
                                        title = result.get('metadata', {}).get('title', '无标题')
                                        text_preview = result.get('text', '')[:100] + ('...' if len(result.get('text', '')) > 100 else '')
                                        results_summary += f"""
                                        <tr>
                                            <td style="border: 1px solid #ddd; padding: 8px;">{idx}</td>
                                            <td style="border: 1px solid #ddd; padding: 8px;">{title}</td>
                                            <td style="border: 1px solid #ddd; padding: 8px;">{text_preview}</td>
                                        </tr>
                                        """
                                    
                                    html_content = f"""
                                    <div style="font-family: Arial, sans-serif; max-width: 800px;">
                                        <h2 style="color: #1677FF;">搜索结果通知</h2>
                                        <p>您好，</p>
                                        <p>您搜索的关键词：<strong>{query}</strong></p>
                                        <p>找到相关结果：<strong>{total}</strong> 条</p>
                                        <p>搜索来源：<strong>{source_info}</strong></p>
                                        <p>搜索时间：{current_time}</p>
                                        <h3>搜索结果摘要（前{min(10, len(paginated_results))}条）：</h3>
                                        <table style="width: 100%; border-collapse: collapse; margin-top: 10px;">
                                            <thead>
                                                <tr style="background-color: #f5f5f5;">
                                                    <th style="border: 1px solid #ddd; padding: 8px; text-align: left;">序号</th>
                                                    <th style="border: 1px solid #ddd; padding: 8px; text-align: left;">标题</th>
                                                    <th style="border: 1px solid #ddd; padding: 8px; text-align: left;">内容摘要</th>
                                                </tr>
                                            </thead>
                                            <tbody>
                                                {results_summary}
                                            </tbody>
                                        </table>
                                        <p style="margin-top: 20px; color: #666;">请登录系统查看完整搜索结果。</p>
                                    </div>
                                    """
                                    if send_notification_email(mail, user_email, subject, html_content):
                                        email_sent = True
                                        logger.info(f"搜索结果邮件已发送到: {user_email}")
                                else:
                                    logger.warning(f"跳过发送邮件到admin邮箱或无效邮箱: {user_email}")
                                except Exception as e:
                                    logger.error(f"发送邮件失败 {user_email}: {e}")
                            else:
                                logger.warning(f"用户邮箱无效或为admin邮箱，跳过发送: {user_email}")
            except Exception as e:
                logger.error(f"发送邮件通知失败: {e}")
        
        return jsonify({
            'results': paginated_results,
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size,
            'email_sent': email_sent if send_email else None,  # 如果用户未选择发送邮件，返回None
            'used_web_search': used_web_search,
            'kb_results_count': len(kb_results),
            'web_results_count': len(web_results),
            'kb_results_by_source': kb_results_by_source,  # 每个知识库的结果数量
            'searched_kbs': kbs_to_search  # 实际搜索的知识库列表
        })
    except Exception as e:
        logger.error(f"搜索失败: {e}")
        return jsonify({'error': f'搜索失败: {str(e)}'}), 500


@app.route('/api/knowledge/upload', methods=['POST'])
def upload_file():
    """上传文件到知识库"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '没有上传文件'}), 400
        
        file = request.files['file']
        kb_name = request.form.get('kb_name', 'default')  # 知识库名称
        
        if file.filename == '':
            return jsonify({'error': '文件名为空'}), 400
        
        # 检查文件大小（限制100MB）
        MAX_FILE_SIZE = 100 * 1024 * 1024  # 100MB
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        if file_size > MAX_FILE_SIZE:
            return jsonify({'error': f'文件太大，最大支持100MB，当前文件大小：{file_size / 1024 / 1024:.2f}MB'}), 400
        
        # 检查知识库总大小（限制500MB）
        MAX_KB_SIZE = 500 * 1024 * 1024  # 500MB
        kb_dir = file_processor.upload_dir / kb_name
        if kb_dir.exists():
            existing_files = list(kb_dir.glob('*'))
            existing_files = [f for f in existing_files if f.is_file()]
            current_kb_size = sum(f.stat().st_size for f in existing_files)
            if current_kb_size + file_size > MAX_KB_SIZE:
                return jsonify({
                    'error': f'知识库总大小超过限制，最大支持500MB，当前：{current_kb_size / 1024 / 1024:.2f}MB，上传后将达到：{(current_kb_size + file_size) / 1024 / 1024:.2f}MB'
                }), 400
        
        # 保存文件到指定知识库文件夹
        # 先获取原始文件名和扩展名
        original_filename = file.filename
        filename = secure_filename(original_filename)
        
        # 如果secure_filename移除了扩展名，尝试从原始文件名恢复
        original_path = Path(original_filename)
        if not Path(filename).suffix and original_path.suffix:
            # 扩展名被移除了，尝试恢复
            safe_name = secure_filename(original_path.stem)
            filename = f"{safe_name}{original_path.suffix}"
        
        kb_dir = file_processor.upload_dir / kb_name
        kb_dir.mkdir(parents=True, exist_ok=True)
        
        file_path = kb_dir / filename
        # 如果文件已存在，添加时间戳
        if file_path.exists():
            name_part = file_path.stem
            ext_part = file_path.suffix
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{name_part}_{timestamp}{ext_part}"
            file_path = kb_dir / filename
        
        file.save(str(file_path))
        
        # 检查文件是否为空
        if file_path.stat().st_size == 0:
            file_path.unlink()  # 删除空文件
            return jsonify({'error': '文件为空，无法处理'}), 400
        
        # 处理文件
        logger.info(f"开始处理文件: {filename}, 路径: {file_path}, 扩展名: {file_path.suffix}")
        texts, metadata_list = file_processor.process_file(file_path, filename)
        
        if not texts:
            logger.warning(f"文件处理失败: {filename}, 扩展名: {file_path.suffix}, 文件大小: {file_path.stat().st_size}")
            return jsonify({'error': f'文件处理失败或文件为空。文件类型: {file_path.suffix or "未知"}，请确保文件格式正确（支持 .txt, .md, .xlsx, .xls, .csv）'}), 400
        
        # 限制单次最多1000条
        texts = texts[:1000]
        if metadata_list:
            metadata_list = metadata_list[:1000]
        
        # 添加到知识库（使用指定的知识库）
        # 注意：这里需要确保知识库实例正确初始化
        kb_instance = KnowledgeBase(db_path=f'instance/faiss_index_{kb_name}')
        kb_instance.add_documents(texts, metadata_list)
        
        # 关闭知识库实例（避免资源占用）
        del kb_instance
        
        # 文件上传不需要发送邮件（根据需求，只在搜索到结果时发送邮件）
        
        return jsonify({
            'message': '文件上传并处理成功',
            'filename': filename,
            'kb_name': kb_name,
            'count': len(texts)
        })
    except Exception as e:
        logger.error(f"文件上传失败: {e}")
        return jsonify({'error': f'上传失败: {str(e)}'}), 500


@app.route('/api/knowledge/add', methods=['POST'])
def add_to_knowledge():
    """添加数据到知识库（JSON格式）"""
    try:
        data = request.get_json()
        texts = data.get('texts', [])
        metadata_list = data.get('metadata', [])
        
        if not texts:
            return jsonify({'error': '数据不能为空'}), 400
        
        # 限制单次最多100条
        texts = texts[:100]
        if metadata_list:
            metadata_list = metadata_list[:100]
        else:
            metadata_list = [{}] * len(texts)
        
        # 添加到default知识库（保持兼容性）
        default_kb = KnowledgeBase(db_path='instance/faiss_index_default')
        default_kb.add_documents(texts, metadata_list)
        del default_kb
        
        # 同时添加到全局kb实例
        kb.add_documents(texts, metadata_list)
        
        # 数据入库不需要发送邮件（根据需求，只在搜索到结果时发送邮件）
        
        return jsonify({
            'message': '数据添加成功',
            'count': len(texts)
        })
    except Exception as e:
        logger.error(f"添加数据失败: {e}")
        return jsonify({'error': f'添加失败: {str(e)}'}), 500


@app.route('/api/knowledge/export', methods=['POST'])
def export_to_excel():
    """导出搜索结果到Excel"""
    try:
        data = request.get_json()
        query = data.get('query', '')
        results = data.get('results', [])
        
        if not results:
            return jsonify({'error': '没有可导出的数据'}), 400
        
        # 创建Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "搜索结果"
        
        # 表头
        headers = ['序号', '标题', '内容', '来源', '链接', '相似度', '发布时间']
        ws.append(headers)
        
        # 数据
        for idx, result in enumerate(results, 1):
            metadata = result.get('metadata', {})
            ws.append([
                idx,
                metadata.get('title', ''),
                result.get('text', '')[:500],  # 限制内容长度
                metadata.get('source', ''),
                metadata.get('link', ''),
                f"{result.get('similarity', 0):.2%}",
                metadata.get('published', '')
            ])
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"搜索结果_{query}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f"导出Excel失败: {e}")
        return jsonify({'error': f'导出失败: {str(e)}'}), 500


@app.route('/api/knowledge/stats', methods=['GET'])
def get_knowledge_stats():
    """获取知识库统计信息（汇总所有知识库）"""
    try:
        total_docs = 0
        total_index_size = 0
        
        # 统计default知识库
        try:
            default_kb = KnowledgeBase(db_path='instance/faiss_index_default')
            default_stats = default_kb.get_stats()
            total_docs += default_stats.get('total_documents', 0)
            total_index_size += default_stats.get('index_size', 0)
            del default_kb
        except:
            pass
        
        # 统计用户创建的知识库
        uploads_dir = file_processor.upload_dir
        if uploads_dir.exists():
            for item in uploads_dir.iterdir():
                if item.is_dir() and item.name != 'default':
                    try:
                        user_kb = KnowledgeBase(db_path=f'instance/faiss_index_{item.name}')
                        user_stats = user_kb.get_stats()
                        total_docs += user_stats.get('total_documents', 0)
                        total_index_size += user_stats.get('index_size', 0)
                        del user_kb
                    except:
                        pass
        
        return jsonify({
            'total_documents': total_docs,
            'index_size': total_index_size
        })
    except Exception as e:
        logger.error(f"获取统计失败: {e}")
        return jsonify({'error': f'获取统计失败: {str(e)}'}), 500


@app.route('/api/knowledge/files', methods=['GET'])
def get_uploaded_files():
    """获取uploads文件夹中的文件列表"""
    try:
        kb_name = request.args.get('kb_name', 'default')
        kb_dir = file_processor.upload_dir / kb_name
        
        files = []
        if kb_dir.exists():
            # 确保只获取文件，不包括子目录
            for file_path in kb_dir.iterdir():
                if file_path.is_file():
                    stat = file_path.stat()
                    # 从实际路径获取知识库名称，确保准确性
                    actual_kb_name = file_path.parent.name
                    
                    # 从数据库读取元数据（如果存在）
                    from models import FileMetadata
                    tags = ''
                    source = ''
                    metadata_record = FileMetadata.query.filter_by(
                        kb_name=actual_kb_name,
                        filename=file_path.name
                    ).first()
                    if metadata_record:
                        tags = metadata_record.tags or ''
                        source = metadata_record.source or ''
                    
                    # 如果数据库中没有，尝试读取旧的元数据文件（兼容性）
                    if not tags and not source:
                        metadata_file = kb_dir / f'.{file_path.name}.metadata'
                        if metadata_file.exists():
                            try:
                                import json
                                with open(metadata_file, 'r', encoding='utf-8') as f:
                                    metadata = json.load(f)
                                    tags = metadata.get('tags', '')
                                    source = metadata.get('source', '')
                                # 迁移到数据库
                                if tags or source:
                                    metadata_record = FileMetadata(
                                        kb_name=actual_kb_name,
                                        filename=file_path.name,
                                        tags=tags,
                                        source=source
                                    )
                                    db.session.add(metadata_record)
                                    db.session.commit()
                                    # 删除旧文件
                                    try:
                                        metadata_file.unlink()
                                    except:
                                        pass
                            except Exception as e:
                                logger.warning(f"读取元数据文件失败 {metadata_file}: {e}")
                    
                    files.append({
                        'filename': file_path.name,
                        'size': stat.st_size,
                        'size_mb': round(stat.st_size / 1024 / 1024, 2),
                        'created': datetime.fromtimestamp(stat.st_ctime).isoformat(),
                        'modified': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                        'kb_name': actual_kb_name,  # 使用实际路径中的知识库名称
                        'tags': tags,
                        'source': source
                    })
        
        # 按修改时间排序
        files.sort(key=lambda x: x['modified'], reverse=True)
        
        return jsonify({
            'files': files,
            'kb_name': kb_name,
            'total': len(files)
        })
    except Exception as e:
        logger.error(f"获取文件列表失败: {e}")
        return jsonify({'error': f'获取文件列表失败: {str(e)}'}), 500


@app.route('/api/knowledge/files/<kb_name>/<filename>', methods=['DELETE'])
def delete_file(kb_name, filename):
    """删除文件"""
    try:
        from urllib.parse import unquote
        filename = unquote(filename)
        
        # 先尝试直接使用文件名
        kb_dir = file_processor.upload_dir / kb_name
        file_path = kb_dir / filename
        
        # 如果文件不存在，尝试使用secure_filename
        if not file_path.exists():
            safe_filename = secure_filename(filename)
            file_path = kb_dir / safe_filename
        
        # 如果还是不存在，尝试模糊匹配
        if not file_path.exists() and kb_dir.exists():
            for f in kb_dir.iterdir():
                if f.is_file() and (f.name == filename or f.name.startswith(filename.split('.')[0])):
                    file_path = f
                    filename = f.name
                    break
        
        if not file_path.exists():
            return jsonify({'error': '文件不存在'}), 404
        
        # 记录文件大小（用于后续统计）
        file_size = file_path.stat().st_size
        
        # 删除文件
        file_path.unlink()
        
        # 验证文件确实被删除了
        if file_path.exists():
            logger.error(f"文件删除失败，文件仍存在: {file_path}")
            return jsonify({'error': '文件删除失败'}), 500
        
        # 注意：从向量数据库中删除数据比较复杂，这里只删除文件
        # 实际应用中，可以通过重建索引或标记删除来处理
        logger.info(f"文件已删除: {file_path}, 大小: {file_size / 1024 / 1024:.2f}MB")
        
        return jsonify({
            'message': '文件删除成功',
            'filename': filename,
            'file_deleted': True  # 标记文件已删除
        })
    except Exception as e:
        logger.error(f"删除文件失败: {e}", exc_info=True)
        return jsonify({'error': f'删除失败: {str(e)}'}), 500


@app.route('/api/knowledge/files/<kb_name>/<filename>/move', methods=['POST'])
def move_file(kb_name, filename):
    """移动文件到其他知识库"""
    try:
        from urllib.parse import unquote
        filename = unquote(filename)
        
        data = request.get_json()
        target_kb = data.get('target_kb')
        
        if not target_kb:
            return jsonify({'error': '目标知识库不能为空'}), 400
        
        if target_kb == kb_name:
            return jsonify({'error': '目标知识库不能与源知识库相同'}), 400
        
        # 查找源文件
        source_kb_dir = file_processor.upload_dir / kb_name
        source_file_path = source_kb_dir / filename
        
        # 如果文件不存在，尝试使用secure_filename
        if not source_file_path.exists():
            safe_filename = secure_filename(filename)
            source_file_path = source_kb_dir / safe_filename
        
        if not source_file_path.exists():
            # 尝试模糊匹配
            if source_kb_dir.exists():
                for f in source_kb_dir.iterdir():
                    if f.is_file() and (f.name == filename or f.name.startswith(filename.split('.')[0])):
                        source_file_path = f
                        filename = f.name
                        break
        
        if not source_file_path.exists():
            return jsonify({'error': '文件不存在'}), 404
        
        # 检查目标知识库是否存在，不存在则创建
        target_kb_dir = file_processor.upload_dir / target_kb
        target_kb_dir.mkdir(parents=True, exist_ok=True)
        
        # 检查目标知识库大小限制
        MAX_KB_SIZE = 500 * 1024 * 1024  # 500MB
        existing_files = list(target_kb_dir.glob('*'))
        existing_files = [f for f in existing_files if f.is_file()]
        current_kb_size = sum(f.stat().st_size for f in existing_files)
        file_size = source_file_path.stat().st_size
        
        if current_kb_size + file_size > MAX_KB_SIZE:
            return jsonify({
                'error': f'目标知识库总大小超过限制，最大支持500MB，当前：{current_kb_size / 1024 / 1024:.2f}MB，移动后将达到：{(current_kb_size + file_size) / 1024 / 1024:.2f}MB'
            }), 400
        
        # 目标文件路径
        target_file_path = target_kb_dir / filename
        # 如果目标文件已存在，添加时间戳
        if target_file_path.exists():
            name_part = target_file_path.stem
            ext_part = target_file_path.suffix
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{name_part}_{timestamp}{ext_part}"
            target_file_path = target_kb_dir / filename
        
        # 移动文件
        import shutil
        shutil.move(str(source_file_path), str(target_file_path))
        logger.info(f"文件已移动: {source_file_path} -> {target_file_path}")
        
        # 验证文件确实被移动了
        if source_file_path.exists():
            logger.error(f"文件移动失败，源文件仍存在: {source_file_path}")
            return jsonify({'error': '文件移动失败'}), 500
        
        if not target_file_path.exists():
            logger.error(f"文件移动失败，目标文件不存在: {target_file_path}")
            return jsonify({'error': '文件移动失败'}), 500
        
        # 处理向量数据库：从源知识库删除，添加到目标知识库
        # 先返回成功，然后异步处理向量数据库，避免超时
        import threading
        
        def process_vector_db():
            """异步处理向量数据库"""
            try:
                logger.info(f"开始处理文件向量数据库: {filename}")
                # 读取文件内容并处理
                texts, metadata_list = file_processor.process_file(target_file_path, filename)
                
                if texts:
                    # 添加到目标知识库
                    target_kb_instance = KnowledgeBase(db_path=f'instance/faiss_index_{target_kb}')
                    target_kb_instance.add_documents(texts, metadata_list)
                    del target_kb_instance
                    
                    logger.info(f"文件已添加到目标知识库向量数据库: {target_kb}, {len(texts)}条数据")
                else:
                    logger.warning(f"文件处理后没有提取到文本: {filename}")
            except Exception as e:
                logger.error(f"更新向量数据库失败: {e}", exc_info=True)
        
        # 启动异步处理线程
        thread = threading.Thread(target=process_vector_db, daemon=True)
        thread.start()
        
        # 注意：从源知识库删除向量数据比较复杂，这里只移动文件
        # 实际应用中，可以通过重建索引或标记删除来处理
        
        return jsonify({
            'message': '文件移动成功，正在后台处理向量数据库',
            'filename': filename,
            'source_kb': kb_name,
            'target_kb': target_kb,
            'file_moved': True  # 标记文件已移动
        })
    except Exception as e:
        logger.error(f"移动文件失败: {e}", exc_info=True)
        return jsonify({'error': f'移动失败: {str(e)}'}), 500


@app.route('/api/knowledge/files/move-batch', methods=['POST'])
def move_files_batch():
    """批量移动文件"""
    try:
        data = request.get_json()
        files = data.get('files', [])  # [{kb_name, filename}, ...]
        target_kb = data.get('target_kb')
        
        if not target_kb:
            return jsonify({'error': '目标知识库不能为空'}), 400
        
        if not files:
            return jsonify({'error': '请选择要移动的文件'}), 400
        
        # 检查目标知识库大小
        target_kb_dir = file_processor.upload_dir / target_kb
        target_kb_dir.mkdir(parents=True, exist_ok=True)
        
        MAX_KB_SIZE = 500 * 1024 * 1024  # 500MB
        existing_files = list(target_kb_dir.glob('*'))
        existing_files = [f for f in existing_files if f.is_file()]
        current_kb_size = sum(f.stat().st_size for f in existing_files)
        
        total_size = 0
        files_to_move = []
        
        # 验证所有文件并计算总大小
        for file_info in files:
            source_kb = file_info.get('kb_name')
            filename = file_info.get('filename')
            
            if source_kb == target_kb:
                continue  # 跳过相同知识库的文件
            
            source_kb_dir = file_processor.upload_dir / source_kb
            source_file_path = source_kb_dir / filename
            
            if not source_file_path.exists():
                safe_filename = secure_filename(filename)
                source_file_path = source_kb_dir / safe_filename
            
            if not source_file_path.exists():
                continue  # 跳过不存在的文件
            
            file_size = source_file_path.stat().st_size
            total_size += file_size
            files_to_move.append({
                'source_kb': source_kb,
                'filename': filename,
                'file_path': source_file_path,
                'size': file_size
            })
        
        if current_kb_size + total_size > MAX_KB_SIZE:
            return jsonify({
                'error': f'目标知识库总大小超过限制，最大支持500MB，当前：{current_kb_size / 1024 / 1024:.2f}MB，移动后将达到：{(current_kb_size + total_size) / 1024 / 1024:.2f}MB'
            }), 400
        
        # 移动所有文件
        moved_files = []
        failed_files = []
        import shutil
        
        for file_info in files_to_move:
            try:
                source_file_path = file_info['file_path']
                filename = file_info['filename']
                target_file_path = target_kb_dir / filename
                
                # 如果目标文件已存在，添加时间戳
                if target_file_path.exists():
                    name_part = target_file_path.stem
                    ext_part = target_file_path.suffix
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"{name_part}_{timestamp}{ext_part}"
                    target_file_path = target_kb_dir / filename
                
                # 移动文件
                shutil.move(str(source_file_path), str(target_file_path))
                
                # 验证文件确实被移动了
                if source_file_path.exists():
                    logger.error(f"文件移动失败，源文件仍存在: {source_file_path}")
                    failed_files.append(file_info['filename'])
                    continue
                
                if not target_file_path.exists():
                    logger.error(f"文件移动失败，目标文件不存在: {target_file_path}")
                    failed_files.append(file_info['filename'])
                    continue
                
                # 处理向量数据库（异步处理，避免超时）
                def process_vector_db_async(file_path, file_name, kb_name):
                    try:
                        logger.info(f"开始处理文件向量数据库: {file_name}")
                        texts, metadata_list = file_processor.process_file(file_path, file_name)
                        if texts:
                            kb_instance = KnowledgeBase(db_path=f'instance/faiss_index_{kb_name}')
                            kb_instance.add_documents(texts, metadata_list)
                            del kb_instance
                            logger.info(f"文件已添加到知识库向量数据库: {kb_name}, {len(texts)}条数据")
                    except Exception as e:
                        logger.error(f"更新向量数据库失败: {e}", exc_info=True)
                
                # 启动异步处理线程
                import threading
                thread = threading.Thread(
                    target=process_vector_db_async,
                    args=(target_file_path, filename, target_kb),
                    daemon=True
                )
                thread.start()
                
                moved_files.append(filename)
                logger.info(f"文件已移动: {source_file_path} -> {target_file_path}")
            except Exception as e:
                logger.error(f"移动文件失败 {file_info['filename']}: {e}")
                failed_files.append(file_info['filename'])
        
        return jsonify({
            'message': f'成功移动 {len(moved_files)} 个文件',
            'moved_files': moved_files,
            'failed_files': failed_files,
            'target_kb': target_kb
        })
    except Exception as e:
        logger.error(f"批量移动文件失败: {e}", exc_info=True)
        return jsonify({'error': f'批量移动失败: {str(e)}'}), 500


@app.route('/api/knowledge/files/<kb_name>/<filename>', methods=['GET'])
def view_file(kb_name, filename):
    """查看文件内容或下载文件"""
    try:
        from urllib.parse import unquote
        filename = unquote(filename)
        
        # 检查是否是下载请求
        download = request.args.get('download', '0') == '1'
        
        # 先尝试直接使用文件名，如果不存在再使用secure_filename
        kb_dir = file_processor.upload_dir / kb_name
        file_path = kb_dir / filename
        
        # 如果文件不存在，尝试使用secure_filename处理后的文件名
        if not file_path.exists():
            safe_filename = secure_filename(filename)
            file_path = kb_dir / safe_filename
        
        if not file_path.exists():
            # 尝试查找所有文件，匹配文件名（忽略扩展名处理）
            if kb_dir.exists():
                for f in kb_dir.iterdir():
                    if f.is_file() and (f.name == filename or f.name.startswith(filename.split('.')[0])):
                        file_path = f
                        filename = f.name
                        break
        
        if not file_path.exists():
            logger.error(f"文件不存在: {kb_dir} / {filename}")
            return jsonify({'error': '文件不存在'}), 404
        
        # 如果是下载请求，直接返回文件
        if download:
            return send_file(
                str(file_path),
                as_attachment=True,
                download_name=filename
            )
        
        # 读取文件内容（限制大小，50MB以下可以预览）
        MAX_PREVIEW_SIZE = 50 * 1024 * 1024  # 50MB
        file_size = file_path.stat().st_size
        
        # 根据文件类型读取内容
        suffix = file_path.suffix.lower()
        
        if suffix in ['.txt', '.md']:
            # 文本文件：尝试多种编码，支持大文件分块读取
            encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1']
            content = None
            used_encoding = None
            
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                        if file_size > 10 * 1024 * 1024:  # 大于10MB，只读取前10MB
                            content = f.read(10 * 1024 * 1024)
                            content += f"\n\n... (文件过大，仅显示前10MB，完整文件大小: {file_size / 1024 / 1024:.2f}MB)"
                        else:
                            content = f.read()
                    used_encoding = encoding
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue
            
            if content is None:
                # 如果所有编码都失败，尝试二进制读取前1MB
                try:
                    with open(file_path, 'rb') as f:
                        raw_content = f.read(1024 * 1024)
                        content = raw_content.decode('utf-8', errors='ignore')
                        if file_size > 1024 * 1024:
                            content += f"\n\n... (文件过大，仅显示前1MB，完整文件大小: {file_size / 1024 / 1024:.2f}MB)"
                except:
                    content = "无法读取文件内容（可能是二进制文件）"
            
            return jsonify({
                'filename': filename,
                'content': content,
                'size': file_size,
                'type': 'text',
                'encoding': used_encoding,
                'is_truncated': file_size > 10 * 1024 * 1024
            })
            
        elif suffix in ['.xlsx', '.xls']:
            # Excel文件返回基本信息
            try:
                wb = load_workbook(file_path, read_only=True)
                sheets_info = []
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    sheets_info.append({
                        'name': sheet_name,
                        'rows': sheet.max_row,
                        'cols': sheet.max_column
                    })
                content = f"Excel文件，包含 {len(wb.sheetnames)} 个工作表"
                return jsonify({
                    'filename': filename,
                    'type': 'excel',
                    'content': content,
                    'sheets': sheets_info,
                    'size': file_size
                })
            except Exception as e:
                logger.error(f"读取Excel文件失败: {e}")
                return jsonify({
                    'filename': filename,
                    'type': 'excel',
                    'content': f"Excel文件（无法读取详细信息: {str(e)}）",
                    'size': file_size
                })
                
        elif suffix == '.csv':
            # CSV文件：读取前1000行
            import csv
            encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1']
            content = None
            rows_read = 0
            
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                        reader = csv.reader(f)
                        rows = []
                        for i, row in enumerate(reader):
                            if i >= 1000:
                                break
                            rows.append(row)
                            rows_read = i + 1
                        content = '\n'.join([','.join([str(cell) for cell in row]) for row in rows])
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue
            
            if content is None:
                content = "无法读取CSV文件内容"
            
            if rows_read >= 1000:
                content += f"\n\n... (仅显示前1000行，文件可能包含更多数据)"
            
            return jsonify({
                'filename': filename,
                'content': content,
                'size': file_size,
                'type': 'csv',
                'rows_displayed': rows_read
            })
            
        elif suffix == '.pdf':
            # PDF文件：返回基本信息
            return jsonify({
                'filename': filename,
                'type': 'pdf',
                'content': f"PDF文件（大小: {file_size / 1024 / 1024:.2f}MB）\n\nPDF文件需要专门的阅读器打开，无法在此预览。",
                'size': file_size,
                'download_url': f'/api/knowledge/files/{kb_name}/{filename}?download=1'
            })
            
        else:
            # 其他文件类型：尝试作为文本读取前10MB
            try:
                encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1']
                content = None
                for encoding in encodings:
                    try:
                        with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                            if file_size > 10 * 1024 * 1024:
                                content = f.read(10 * 1024 * 1024)
                                content += f"\n\n... (文件过大，仅显示前10MB，完整文件大小: {file_size / 1024 / 1024:.2f}MB)"
                            else:
                                content = f.read()
                        break
                    except (UnicodeDecodeError, UnicodeError):
                        continue
                
                if content is None:
                    content = f"无法预览此文件类型 ({suffix})，文件大小: {file_size / 1024 / 1024:.2f}MB"
                    
                return jsonify({
                    'filename': filename,
                    'content': content,
                    'size': file_size,
                    'type': suffix[1:] if suffix else 'unknown',
                    'is_truncated': file_size > 10 * 1024 * 1024
                })
            except Exception as e:
                logger.error(f"读取文件失败: {e}")
                return jsonify({
                    'filename': filename,
                    'type': suffix[1:] if suffix else 'unknown',
                    'content': f"无法预览此文件类型 ({suffix})，错误: {str(e)}",
                    'size': file_size
                })
                
    except Exception as e:
        logger.error(f"查看文件失败: {e}", exc_info=True)
        return jsonify({'error': f'查看文件失败: {str(e)}'}), 500


@app.route('/api/knowledge/kb-list', methods=['GET'])
def get_kb_list():
    """获取所有知识库列表"""
    try:
        kb_list = []
        uploads_dir = file_processor.upload_dir
        
        if uploads_dir.exists():
            for item in uploads_dir.iterdir():
                if item.is_dir():
                    # 统计文件夹中的文件数量和大小（只统计文件，不包括子目录）
                    files = []
                    try:
                        for f in item.iterdir():
                            if f.is_file():
                                files.append(f)
                    except (PermissionError, OSError) as e:
                        logger.warning(f"无法访问知识库目录 {item.name}: {e}")
                        continue
                    
                    total_size = sum(f.stat().st_size for f in files)
                    
                    kb_list.append({
                        'name': item.name,
                        'file_count': len(files),
                        'total_size': total_size,
                        'total_size_mb': round(total_size / 1024 / 1024, 2)
                    })
        
        # 如果没有知识库，添加默认的
        if not kb_list:
            # 确保default目录存在
            default_dir = uploads_dir / 'default'
            default_dir.mkdir(parents=True, exist_ok=True)
            kb_list.append({
                'name': 'default',
                'file_count': 0,
                'total_size': 0,
                'total_size_mb': 0
            })
        
        # 按名称排序
        kb_list.sort(key=lambda x: x['name'])
        
        return jsonify({
            'kb_list': kb_list,
            'total': len(kb_list)
        })
    except Exception as e:
        logger.error(f"获取知识库列表失败: {e}")
        return jsonify({'error': f'获取知识库列表失败: {str(e)}'}), 500


@app.route('/api/knowledge/kb-create', methods=['POST'])
def create_kb():
    """创建新知识库"""
    try:
        data = request.get_json()
        kb_name = data.get('kb_name', '').strip()
        
        if not kb_name:
            return jsonify({'error': '知识库名称不能为空'}), 400
        
        # 验证名称（只允许字母、数字、下划线、中文字符）
        import re
        if not re.match(r'^[\w\u4e00-\u9fa5]+$', kb_name):
            return jsonify({'error': '知识库名称只能包含字母、数字、下划线和中文'}), 400
        
        # 检查是否已存在
        kb_dir = file_processor.upload_dir / kb_name
        if kb_dir.exists():
            return jsonify({'error': '知识库已存在'}), 400
        
        # 创建知识库文件夹
        kb_dir.mkdir(parents=True, exist_ok=True)
        
        # 创建对应的知识库索引（初始化即可）
        kb_instance = KnowledgeBase(db_path=f'instance/faiss_index_{kb_name}')
        # 保存索引以确保创建成功
        kb_instance.save_index()
        del kb_instance
        
        return jsonify({
            'message': '知识库创建成功',
            'kb_name': kb_name
        })
    except Exception as e:
        logger.error(f"创建知识库失败: {e}")
        return jsonify({'error': f'创建失败: {str(e)}'}), 500


@app.route('/api/knowledge/check-email', methods=['POST'])
def check_email_status():
    """检查邮件发送状态"""
    try:
        data = request.get_json()
        user_email = data.get('email')
        
        if not user_email:
            return jsonify({'email_sent': False, 'message': '邮箱地址不能为空'}), 400
        
        # 检查邮件配置
        mail_configured = bool(app.config.get('MAIL_USERNAME') and app.config.get('MAIL_PASSWORD'))
        
        return jsonify({
            'email_sent': mail_configured,
            'mail_configured': mail_configured,
            'message': '邮件服务已配置' if mail_configured else '邮件服务未配置'
        })
    except Exception as e:
        return jsonify({'error': f'检查失败: {str(e)}'}), 500


@app.route('/api/knowledge/send-search-email', methods=['POST'])
def send_search_email():
    """搜索后发送邮件通知"""
    try:
        data = request.get_json()
        query = data.get('query', '')
        total = data.get('total', 0)
        results = data.get('results', [])
        selected_kbs = data.get('selected_kbs', [])
        
        if not query or total == 0:
            return jsonify({'error': '搜索关键词或结果不能为空'}), 400
        
        # 从请求头获取token
        auth_header = request.headers.get('Authorization', '')
        user_email = None
        if auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
            payload = verify_token(token)
            if payload:
                user = User.query.get(payload['user_id'])
                if user and user.is_verified:
                    user_email = user.email
        
        # 过滤掉admin默认邮箱，避免发送到无效邮箱
        ADMIN_EMAILS = ['admin@xu-news.com', 'admin@example.com']
        
        if not user_email:
            return jsonify({'error': '用户未登录或未验证'}), 401
        
        # 检查是否为admin邮箱
        if user_email in ADMIN_EMAILS:
            logger.warning(f"尝试发送邮件到admin邮箱，已阻止: {user_email}")
            return jsonify({'error': '不能发送邮件到系统默认邮箱'}), 400
        
        # 验证邮箱格式
        if '@' not in user_email:
            logger.warning(f"邮箱格式无效: {user_email}")
            return jsonify({'error': '邮箱格式无效'}), 400
        
        # 发送邮件
        email_sent = False
        try:
            source_info = "用户创建的知识库" if selected_kbs else "联网搜索"
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            subject = f"搜索结果通知 - {query[:30]}{'...' if len(query) > 30 else ''} ({total}条结果)"
            
            # 构建结果摘要表格
            results_summary = ""
            for idx, result in enumerate(results[:10], 1):  # 最多显示10条
                title = result.get('metadata', {}).get('title', '无标题')
                text_preview = result.get('text', '')[:100] + ('...' if len(result.get('text', '')) > 100 else '')
                results_summary += f"""
                <tr>
                    <td style="border: 1px solid #ddd; padding: 8px;">{idx}</td>
                    <td style="border: 1px solid #ddd; padding: 8px;">{title}</td>
                    <td style="border: 1px solid #ddd; padding: 8px;">{text_preview}</td>
                </tr>
                """
            
            html_content = f"""
            <div style="font-family: Arial, sans-serif; max-width: 800px;">
                <h2 style="color: #1677FF;">搜索结果通知</h2>
                <p>您好，</p>
                <p>您搜索的关键词：<strong>{query}</strong></p>
                <p>找到相关结果：<strong>{total}</strong> 条</p>
                <p>搜索来源：<strong>{source_info}</strong></p>
                <p>搜索时间：{current_time}</p>
                <h3>搜索结果摘要（前{min(10, len(results))}条）：</h3>
                <table style="width: 100%; border-collapse: collapse; margin-top: 10px;">
                    <thead>
                        <tr style="background-color: #f5f5f5;">
                            <th style="border: 1px solid #ddd; padding: 8px; text-align: left;">序号</th>
                            <th style="border: 1px solid #ddd; padding: 8px; text-align: left;">标题</th>
                            <th style="border: 1px solid #ddd; padding: 8px; text-align: left;">内容摘要</th>
                        </tr>
                    </thead>
                    <tbody>
                        {results_summary}
                    </tbody>
                </table>
                <p style="margin-top: 20px; color: #666;">请登录系统查看完整搜索结果。</p>
            </div>
            """
            if send_notification_email(mail, user_email, subject, html_content):
                email_sent = True
                logger.info(f"搜索结果邮件已发送到: {user_email}")
        except Exception as e:
            logger.error(f"发送邮件失败 {user_email}: {e}")
        
        return jsonify({
            'email_sent': email_sent,
            'message': '邮件发送成功' if email_sent else '邮件发送失败'
        })
    except Exception as e:
        logger.error(f"发送搜索邮件失败: {e}")
        return jsonify({'error': f'发送失败: {str(e)}'}), 500


@app.route('/api/news/crawl', methods=['POST'])
def manual_crawl():
    """手动触发新闻采集"""
    try:
        data = request.get_json()
        rss_urls = data.get('rss_urls', [])
        web_urls = data.get('web_urls', [])
        
        crawler = NewsCrawler()
        articles = crawler.crawl_multiple_sources(rss_urls=rss_urls, web_urls=web_urls)
        
        if articles:
            texts = []
            metadata_list = []
            for article in articles:
                text = f"标题：{article.get('title', '')}\n内容：{article.get('content', '')}"
                texts.append(text)
                metadata_list.append({
                    'title': article.get('title', ''),
                    'source': article.get('source', ''),
                    'link': article.get('link', ''),
                    'published': article.get('published', ''),
                })
            
            # 添加到default知识库（确保手动采集的新闻可以被搜索到）
            default_kb = KnowledgeBase(db_path='instance/faiss_index_default')
            default_kb.add_documents(texts, metadata_list)
            del default_kb
            
            # 同时添加到全局kb实例（保持兼容）
            kb.add_documents(texts, metadata_list)
            
            # 新闻采集不需要发送邮件（根据需求，只在搜索到结果时发送邮件）
        
        return jsonify({
            'message': '采集完成',
            'count': len(articles)
        })
    except Exception as e:
        logger.error(f"手动采集失败: {e}")
        return jsonify({'error': f'采集失败: {str(e)}'}), 500


@app.route('/api/knowledge/files/<kb_name>/<filename>/metadata', methods=['PUT'])
def update_file_metadata(kb_name, filename):
    """更新文件元数据（标签、来源）- 存储在数据库中，不创建文件"""
    try:
        data = request.get_json()
        tags = data.get('tags', '')
        source = data.get('source', '')
        
        # 解码文件名
        filename = filename.replace('%20', ' ')
        
        # 查找文件
        uploads_dir = file_processor.upload_dir
        kb_dir = uploads_dir / kb_name
        if not kb_dir.exists():
            return jsonify({'error': '知识库不存在'}), 404
        
        # 尝试多种方式查找文件
        file_path = None
        for f in kb_dir.iterdir():
            if f.is_file():
                if f.name == filename or secure_filename(f.name) == filename:
                    file_path = f
                    break
        
        if not file_path or not file_path.exists():
            return jsonify({'error': '文件不存在'}), 404
        
        # 删除旧的元数据文件（如果存在）
        metadata_file = kb_dir / f'.{filename}.metadata'
        if metadata_file.exists():
            try:
                metadata_file.unlink()
                logger.info(f"删除旧的元数据文件: {metadata_file}")
            except Exception as e:
                logger.warning(f"删除元数据文件失败: {e}")
        
        # 将元数据存储到数据库
        from models import FileMetadata
        
        # 查找或创建元数据记录
        metadata_record = FileMetadata.query.filter_by(
            kb_name=kb_name,
            filename=filename
        ).first()
        
        if metadata_record:
            # 更新现有记录
            metadata_record.tags = tags
            metadata_record.source = source
            metadata_record.updated_at = datetime.now()
        else:
            # 创建新记录
            metadata_record = FileMetadata(
                kb_name=kb_name,
                filename=filename,
                tags=tags,
                source=source
            )
            db.session.add(metadata_record)
        
        db.session.commit()
        
        logger.info(f"更新文件元数据: {kb_name}/{filename}, tags={tags}, source={source}")
        
        return jsonify({
            'message': '元数据更新成功',
            'tags': tags,
            'source': source
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f"更新元数据失败: {e}")
        return jsonify({'error': f'更新失败: {str(e)}'}), 500


def extract_keywords(text, top_k=10):
    """提取关键词（快速版本）"""
    # 使用jieba分词
    try:
        words = jieba.cut(text)
        # 过滤停用词和单字
        stop_words = {'的', '了', '在', '是', '我', '有', '和', '就', '不', '人', '都', '一', '一个', '上', '也', '很', '到', '说', '要', '去', '你', '会', '着', '没有', '看', '好', '自己', '这'}
        words = [w for w in words if len(w) > 1 and w not in stop_words and w.strip()]
        counter = Counter(words)
        return [{'keyword': word, 'count': count} for word, count in counter.most_common(top_k)]
    except:
        # 如果jieba不可用，使用简单的词频统计
        words = re.findall(r'\w+', text)
        words = [w for w in words if len(w) > 1]
        counter = Counter(words)
        return [{'keyword': word, 'count': count} for word, count in counter.most_common(top_k)]


def calculate_keyword_similarity(keywords1, keywords2):
    """计算两个关键词列表的相似度（Jaccard相似度）"""
    set1 = set(keywords1)
    set2 = set(keywords2)
    if len(set1) == 0 and len(set2) == 0:
        return 1.0
    if len(set1) == 0 or len(set2) == 0:
        return 0.0
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    return intersection / union if union > 0 else 0.0


def generate_cluster_summary(clusters, total_docs, top_keywords):
    """生成聚类分析结论"""
    if not clusters or total_docs == 0:
        return {
            'summary': '数据量不足，无法进行有效的聚类分析。',
            'insights': []
        }
    
    insights = []
    summary_parts = []
    
    # 1. 聚类数量分析
    cluster_count = len(clusters)
    avg_cluster_size = total_docs / cluster_count if cluster_count > 0 else 0
    summary_parts.append(f'共识别出 {cluster_count} 个主要聚类，平均每个聚类包含 {avg_cluster_size:.1f} 条数据。')
    
    # 2. 聚类分布分析
    cluster_sizes = [c['count'] for c in clusters]
    max_cluster = max(cluster_sizes)
    min_cluster = min(cluster_sizes)
    if max_cluster > min_cluster * 2:
        insights.append(f'聚类分布不均匀：最大聚类包含 {max_cluster} 条数据，最小聚类仅 {min_cluster} 条，数据集中度较高。')
    else:
        insights.append('聚类分布相对均匀，各主题内容分布较为平衡。')
    
    # 3. 主要主题识别
    if clusters:
        largest_cluster = clusters[0]
        if largest_cluster['keywords']:
            main_theme = ', '.join(largest_cluster['keywords'][:3])
            summary_parts.append(f'主要主题为：{main_theme}（包含 {largest_cluster["count"]} 条数据，占比 {largest_cluster["count"]/total_docs*100:.1f}%）。')
    
    # 4. 关键词分析
    if top_keywords:
        top_3_keywords = [kw['keyword'] for kw in top_keywords[:3]]
        insights.append(f'核心关键词为：{", ".join(top_3_keywords)}，反映了数据的主要内容特征。')
    
    # 5. 数据质量评估
    if cluster_count >= 3:
        insights.append('聚类结构清晰，数据具有良好的主题区分度。')
    elif cluster_count == 2:
        insights.append('数据主要分为两个主题，建议进一步细分以提高分析精度。')
    else:
        insights.append('数据主题较为单一，建议增加数据多样性。')
    
    # 6. 聚类特征分析
    unique_themes = set()
    for cluster in clusters:
        if cluster.get('keywords'):
            theme = ', '.join(cluster['keywords'][:2])
            unique_themes.add(theme)
    
    if len(unique_themes) == cluster_count:
        insights.append('各聚类主题特征明显，关键词区分度高。')
    else:
        insights.append('部分聚类主题存在重叠，建议调整聚类参数以提高区分度。')
    
    summary = ' '.join(summary_parts)
    
    return {
        'summary': summary,
        'insights': insights
    }


def fast_cluster_documents(documents, progress_callback=None):
    """改进的快速聚类算法（基于关键词相似度）"""
    if not documents or len(documents) == 0:
        return []
    
    total = len(documents)
    clusters = []
    
    # 第一步：提取每个文档的关键词（30%时间）
    if progress_callback:
        progress_callback(0.1, '提取文档关键词...')
    
    doc_keywords = []
    for i, doc in enumerate(documents):
        text = doc.get('text', '')
        keywords = extract_keywords(text, top_k=8)  # 增加关键词数量以提高区分度
        doc_keywords.append([kw['keyword'] for kw in keywords])
        if progress_callback and i % max(1, total // 10) == 0:
            progress_callback(0.1 + (i / total) * 0.3, f'处理文档 {i+1}/{total}...')
    
    # 第二步：基于关键词相似度的真实聚类（60%时间）
    if progress_callback:
        progress_callback(0.4, '执行聚类分析...')
    
    # 使用改进的聚类算法：基于关键词相似度分组
    # 目标聚类数：根据文档数量动态调整（3-8个）
    target_clusters = min(8, max(3, total // 100 + 3))
    similarity_threshold = 0.3  # 相似度阈值
    
    # 初始化：每个文档一个聚类
    doc_clusters = [None] * total  # None表示未分配
    cluster_centers = []  # 每个聚类的中心（关键词集合）
    cluster_docs = []  # 每个聚类包含的文档索引
    
    # 为每个文档分配聚类
    for i in range(total):
        if doc_clusters[i] is not None:
            continue
        
        # 尝试找到相似的现有聚类
        best_cluster = None
        best_similarity = 0
        
        for cluster_idx, center_keywords in enumerate(cluster_centers):
            similarity = calculate_keyword_similarity(doc_keywords[i], center_keywords)
            if similarity > best_similarity and similarity >= similarity_threshold:
                best_similarity = similarity
                best_cluster = cluster_idx
        
        if best_cluster is not None:
            # 分配到现有聚类
            doc_clusters[i] = best_cluster
            cluster_docs[best_cluster].append(i)
            # 更新聚类中心（合并关键词）
            cluster_centers[best_cluster] = list(set(cluster_centers[best_cluster] + doc_keywords[i]))[:10]
        else:
            # 创建新聚类
            if len(cluster_centers) < target_clusters:
                new_cluster_idx = len(cluster_centers)
                doc_clusters[i] = new_cluster_idx
                cluster_centers.append(doc_keywords[i][:10])
                cluster_docs.append([i])
            else:
                # 如果已达到最大聚类数，分配到最相似的聚类
                best_cluster = 0
                best_similarity = 0
                for cluster_idx, center_keywords in enumerate(cluster_centers):
                    similarity = calculate_keyword_similarity(doc_keywords[i], center_keywords)
                    if similarity > best_similarity:
                        best_similarity = similarity
                        best_cluster = cluster_idx
                doc_clusters[i] = best_cluster
                cluster_docs[best_cluster].append(i)
                cluster_centers[best_cluster] = list(set(cluster_centers[best_cluster] + doc_keywords[i]))[:10]
        
        if progress_callback and i % max(1, total // 10) == 0:
            progress = 0.4 + (i / total) * 0.5
            progress_callback(progress, f'聚类文档 {i+1}/{total}...')
    
    # 第三步：生成聚类结果（10%时间）
    if progress_callback:
        progress_callback(0.9, '生成聚类结果...')
    
    for cluster_idx, doc_indices in enumerate(cluster_docs):
        if len(doc_indices) == 0:
            continue
        
        # 合并该聚类中所有文档的关键词
        all_cluster_keywords = []
        for doc_idx in doc_indices:
            all_cluster_keywords.extend(doc_keywords[doc_idx])
        
        # 提取聚类的主要关键词（去重并统计）
        cluster_keywords = extract_keywords(' '.join(all_cluster_keywords), top_k=5)
        
        # 计算聚类主题（基于主要关键词）
        main_keywords = [kw['keyword'] for kw in cluster_keywords[:3]]
        
        clusters.append({
            'cluster': f'聚类 {len(clusters) + 1}',
            'count': len(doc_indices),
            'keywords': main_keywords,
            'theme': ', '.join(main_keywords) if main_keywords else '未分类'
        })
    
    # 按文档数量排序（从多到少）
    clusters.sort(key=lambda x: x['count'], reverse=True)
    
    # 重新编号
    for i, cluster in enumerate(clusters):
        cluster['cluster'] = f'聚类 {i + 1}'
    
    if progress_callback:
        progress_callback(1.0, '聚类完成')
    
    return clusters


def run_analysis_task(task_id, task_type, kb_name, filename=None):
    """后台运行分析任务"""
    try:
        with analysis_tasks_lock:
            analysis_tasks[task_id]['status'] = 'running'
            analysis_tasks[task_id]['start_time'] = time.time()
        
        def update_progress(progress, message):
            with analysis_tasks_lock:
                if task_id in analysis_tasks:
                    analysis_tasks[task_id]['progress'] = progress
                    analysis_tasks[task_id]['message'] = message
                    elapsed = time.time() - analysis_tasks[task_id]['start_time']
                    if progress > 0:
                        estimated_total = elapsed / progress
                        remaining = estimated_total - elapsed
                        analysis_tasks[task_id]['estimated_remaining'] = max(0, remaining)
        
        if task_type == 'document':
            # 分析单个文档
            filename = filename.replace('%20', ' ')
            uploads_dir = file_processor.upload_dir
            kb_dir = uploads_dir / kb_name
            
            file_path = None
            for f in kb_dir.iterdir():
                if f.is_file() and not f.name.startswith('.'):
                    if f.name == filename or secure_filename(f.name) == filename:
                        file_path = f
                        break
            
            if not file_path or not file_path.exists():
                with analysis_tasks_lock:
                    analysis_tasks[task_id]['status'] = 'error'
                    analysis_tasks[task_id]['error'] = '文件不存在'
                return
            
            update_progress(0.1, '读取文件...')
            texts, _ = file_processor.process_file(file_path, filename)
            
            update_progress(0.3, '提取关键词...')
            all_text = ' '.join(texts)
            top_keywords = extract_keywords(all_text, top_k=10)
            
            update_progress(0.5, '执行聚类分析...')
            # 将文本块转换为文档格式
            documents = [{'text': text} for text in texts]
            clusters = fast_cluster_documents(documents, update_progress)
            
            # 生成聚类分析结论
            cluster_summary = generate_cluster_summary(clusters, len(texts), top_keywords)
            
            with analysis_tasks_lock:
                analysis_tasks[task_id]['status'] = 'completed'
                analysis_tasks[task_id]['result'] = {
                    'total_documents': len(texts),
                    'top_keywords': top_keywords,
                    'clusters': clusters,
                    'cluster_summary': cluster_summary
                }
                analysis_tasks[task_id]['progress'] = 1.0
                analysis_tasks[task_id]['message'] = '分析完成'
                
        elif task_type == 'knowledge_base':
            # 分析整个知识库
            update_progress(0.1, '加载知识库...')
            if kb_name == 'default':
                kb_instance = KnowledgeBase(db_path='instance/faiss_index_default')
            else:
                kb_instance = KnowledgeBase(db_path=f'instance/faiss_index_{kb_name}')
            
            documents = kb_instance.documents
            if len(documents) == 0:
                with analysis_tasks_lock:
                    analysis_tasks[task_id]['status'] = 'completed'
                    analysis_tasks[task_id]['result'] = {
                        'total_documents': 0,
                        'top_keywords': [],
                        'clusters': []
                    }
                    analysis_tasks[task_id]['progress'] = 1.0
                del kb_instance
                return
            
            update_progress(0.2, '提取关键词...')
            all_text = ' '.join([doc.get('text', '') for doc in documents])
            top_keywords = extract_keywords(all_text, top_k=10)
            
            update_progress(0.4, '执行聚类分析...')
            clusters = fast_cluster_documents(documents, update_progress)
            
            del kb_instance
            
            # 生成聚类分析结论
            cluster_summary = generate_cluster_summary(clusters, len(documents), top_keywords)
            
            with analysis_tasks_lock:
                analysis_tasks[task_id]['status'] = 'completed'
                analysis_tasks[task_id]['result'] = {
                    'total_documents': len(documents),
                    'top_keywords': top_keywords,
                    'clusters': clusters,
                    'cluster_summary': cluster_summary
                }
                analysis_tasks[task_id]['progress'] = 1.0
                analysis_tasks[task_id]['message'] = '分析完成'
                
    except Exception as e:
        logger.error(f"分析任务失败 {task_id}: {e}")
        with analysis_tasks_lock:
            if task_id in analysis_tasks:
                analysis_tasks[task_id]['status'] = 'error'
                analysis_tasks[task_id]['error'] = str(e)


@app.route('/api/knowledge/analyze-document/<kb_name>/<filename>', methods=['POST'])
def analyze_document_async(kb_name, filename):
    """异步分析单个文档（创建任务）"""
    try:
        task_id = str(uuid.uuid4())
        filename_decoded = filename.replace('%20', ' ')
        
        with analysis_tasks_lock:
            analysis_tasks[task_id] = {
                'task_id': task_id,
                'type': 'document',
                'kb_name': kb_name,
                'filename': filename_decoded,
                'status': 'pending',
                'progress': 0.0,
                'message': '等待开始...',
                'estimated_remaining': 0,
                'start_time': None,
                'result': None,
                'error': None
            }
        
        # 在后台线程中运行
        thread = threading.Thread(target=run_analysis_task, args=(task_id, 'document', kb_name, filename_decoded))
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'task_id': task_id,
            'status': 'pending',
            'message': '分析任务已创建'
        })
    except Exception as e:
        logger.error(f"创建分析任务失败: {e}")
        return jsonify({'error': f'创建任务失败: {str(e)}'}), 500


@app.route('/api/knowledge/analyze-kb/<kb_name>', methods=['POST'])
def analyze_knowledge_base_async(kb_name):
    """异步分析整个知识库（创建任务）"""
    try:
        task_id = str(uuid.uuid4())
        
        with analysis_tasks_lock:
            analysis_tasks[task_id] = {
                'task_id': task_id,
                'type': 'knowledge_base',
                'kb_name': kb_name,
                'filename': None,
                'status': 'pending',
                'progress': 0.0,
                'message': '等待开始...',
                'estimated_remaining': 0,
                'start_time': None,
                'result': None,
                'error': None
            }
        
        # 在后台线程中运行
        thread = threading.Thread(target=run_analysis_task, args=(task_id, 'knowledge_base', kb_name, None))
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'task_id': task_id,
            'status': 'pending',
            'message': '分析任务已创建'
        })
    except Exception as e:
        logger.error(f"创建分析任务失败: {e}")
        return jsonify({'error': f'创建任务失败: {str(e)}'}), 500


@app.route('/api/knowledge/analysis-task/<task_id>', methods=['GET'])
def get_analysis_task_status(task_id):
    """获取分析任务状态"""
    try:
        with analysis_tasks_lock:
            if task_id not in analysis_tasks:
                return jsonify({'error': '任务不存在'}), 404
            
            task = analysis_tasks[task_id].copy()
            # 格式化剩余时间
            if task.get('estimated_remaining'):
                remaining_seconds = int(task['estimated_remaining'])
                if remaining_seconds < 60:
                    task['estimated_remaining_text'] = f'{remaining_seconds}秒'
                elif remaining_seconds < 3600:
                    task['estimated_remaining_text'] = f'{remaining_seconds // 60}分{remaining_seconds % 60}秒'
                else:
                    hours = remaining_seconds // 3600
                    minutes = (remaining_seconds % 3600) // 60
                    task['estimated_remaining_text'] = f'{hours}小时{minutes}分钟'
            else:
                task['estimated_remaining_text'] = '计算中...'
            
            return jsonify(task)
    except Exception as e:
        logger.error(f"获取任务状态失败: {e}")
        return jsonify({'error': f'获取状态失败: {str(e)}'}), 500


@app.route('/api/knowledge/analysis-tasks', methods=['GET'])
def list_analysis_tasks():
    """获取所有分析任务列表"""
    try:
        with analysis_tasks_lock:
            tasks = []
            for task_id, task in analysis_tasks.items():
                task_info = {
                    'task_id': task_id,
                    'type': task.get('type'),
                    'kb_name': task.get('kb_name'),
                    'filename': task.get('filename'),
                    'status': task.get('status'),
                    'progress': task.get('progress', 0.0),
                    'message': task.get('message', ''),
                    'estimated_remaining': task.get('estimated_remaining', 0)
                }
                if task_info['estimated_remaining']:
                    remaining_seconds = int(task_info['estimated_remaining'])
                    if remaining_seconds < 60:
                        task_info['estimated_remaining_text'] = f'{remaining_seconds}秒'
                    elif remaining_seconds < 3600:
                        task_info['estimated_remaining_text'] = f'{remaining_seconds // 60}分{remaining_seconds % 60}秒'
                    else:
                        hours = remaining_seconds // 3600
                        minutes = (remaining_seconds % 3600) // 60
                        task_info['estimated_remaining_text'] = f'{hours}小时{minutes}分钟'
                else:
                    task_info['estimated_remaining_text'] = '计算中...'
                tasks.append(task_info)
            
            return jsonify({'tasks': tasks})
    except Exception as e:
        logger.error(f"获取任务列表失败: {e}")
        return jsonify({'error': f'获取列表失败: {str(e)}'}), 500


if __name__ == '__main__':
    app.run(debug=True, port=5000, host='0.0.0.0')

