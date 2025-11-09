"""文件处理模块，支持多种文件格式"""
import os
from pathlib import Path
from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class FileProcessor:
    def __init__(self, upload_dir='uploads'):
        # 确保uploads文件夹在项目根目录
        base_path = Path(__file__).parent.parent
        self.upload_dir = base_path / upload_dir
        self.upload_dir.mkdir(parents=True, exist_ok=True)
    
    def process_file(self, file_path, filename):
        """处理上传的文件，返回文本列表和元数据列表"""
        file_path = Path(file_path)
        suffix = file_path.suffix.lower()
        
        # 如果文件路径没有扩展名，尝试从文件名获取
        if not suffix:
            filename_path = Path(filename)
            suffix = filename_path.suffix.lower()
        
        logger.info(f"处理文件: {filename}, 扩展名: {suffix}, 文件路径: {file_path}")
        
        texts = []
        metadata_list = []
        
        try:
            # 检查文件是否存在
            if not file_path.exists():
                logger.error(f"文件不存在: {file_path}")
                return [], []
            
            # 检查文件大小
            file_size = file_path.stat().st_size
            if file_size == 0:
                logger.warning(f"文件为空: {file_path}")
                return [], []
            
            if suffix == '.txt' or suffix == '.md':
                # 文本文件
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    # 按段落分割
                    paragraphs = [p.strip() for p in content.split('\n\n') if p.strip()]
                    for para in paragraphs:
                        texts.append(para)
                        metadata_list.append({
                            'title': filename,
                            'source': '文件上传',
                            'file_name': filename,
                            'file_type': suffix[1:],
                        })
            
            elif suffix in ['.xlsx', '.xls']:
                # Excel文件
                wb = load_workbook(file_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    # 读取所有单元格内容
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                        row_text = []
                        for col_idx, cell_value in enumerate(row, 1):
                            if cell_value:
                                row_text.append(str(cell_value))
                        
                        if row_text:
                            text = ' '.join(row_text)
                            texts.append(text)
                            metadata_list.append({
                                'title': f'{filename} - {sheet_name} - 行{row_idx}',
                                'source': '文件上传',
                                'file_name': filename,
                                'file_type': 'excel',
                                'sheet': sheet_name,
                                'row': row_idx,
                            })
            
            elif suffix == '.csv':
                # CSV文件
                import csv
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for row_idx, row in enumerate(reader, 1):
                        if row:
                            text = ' '.join([str(cell) for cell in row if cell])
                            texts.append(text)
                            metadata_list.append({
                                'title': f'{filename} - 行{row_idx}',
                                'source': '文件上传',
                                'file_name': filename,
                                'file_type': 'csv',
                                'row': row_idx,
                            })
            
            elif suffix == '.pdf':
                # PDF文件
                try:
                    # 尝试使用PyPDF2
                    try:
                        import PyPDF2
                        with open(file_path, 'rb') as f:
                            pdf_reader = PyPDF2.PdfReader(f)
                            logger.info(f"PDF文件共有 {len(pdf_reader.pages)} 页")
                            
                            # 检查是否加密
                            if pdf_reader.is_encrypted:
                                logger.warning(f"PDF文件已加密，尝试解密: {filename}")
                                try:
                                    pdf_reader.decrypt('')  # 尝试空密码解密
                                except:
                                    logger.error(f"无法解密PDF文件: {filename}")
                                    return [], []
                            
                            for page_idx, page in enumerate(pdf_reader.pages, 1):
                                try:
                                    page_text = page.extract_text()
                                    # 检查是否成功提取文本
                                    if page_text is None:
                                        logger.debug(f"PDF第{page_idx}页 extract_text() 返回 None")
                                        continue
                                    
                                    # 清理文本（移除多余空白）
                                    page_text = page_text.strip()
                                    if not page_text:
                                        logger.debug(f"PDF第{page_idx}页文本为空")
                                        continue
                                    
                                    # 清理文本（移除多余空白）
                                    page_text = ' '.join(page_text.split())
                                    
                                    # 按段落分割
                                    paragraphs = [p.strip() for p in page_text.split('\n\n') if p.strip()]
                                    if not paragraphs:
                                        paragraphs = [p.strip() for p in page_text.split('\n') if p.strip()]
                                    
                                    # 如果段落太多，合并成较大的块
                                    if len(paragraphs) > 100:
                                        # 合并段落，每500字符一个块
                                        merged_text = ' '.join(paragraphs)
                                        chunk_size = 500
                                        for i in range(0, len(merged_text), chunk_size):
                                            chunk = merged_text[i:i+chunk_size].strip()
                                            if chunk and len(chunk) > 10:
                                                texts.append(chunk)
                                                metadata_list.append({
                                                    'title': f'{filename} - 第{page_idx}页',
                                                    'source': '文件上传',
                                                    'file_name': filename,
                                                    'file_type': 'pdf',
                                                    'page': page_idx,
                                                })
                                    else:
                                        for para in paragraphs:
                                            if para and len(para) > 10:  # 过滤太短的段落
                                                texts.append(para)
                                                metadata_list.append({
                                                    'title': f'{filename} - 第{page_idx}页',
                                                    'source': '文件上传',
                                                    'file_name': filename,
                                                    'file_type': 'pdf',
                                                    'page': page_idx,
                                                })
                                    
                                    if page_idx == 1 and texts:
                                        logger.info(f"PDF第1页成功提取文本，长度: {len(page_text)}")
                                        
                                except Exception as e:
                                    logger.warning(f"提取PDF第{page_idx}页失败: {e}")
                                    continue
                    except ImportError:
                        logger.error("PDF处理需要安装 PyPDF2: pip install PyPDF2")
                        return [], []
                    except Exception as e:
                        logger.error(f"PyPDF2处理失败，尝试使用pdfplumber: {e}")
                        # 如果PyPDF2失败，尝试使用pdfplumber
                        try:
                            import pdfplumber
                            with pdfplumber.open(file_path) as pdf:
                                logger.info(f"使用pdfplumber处理PDF，共有 {len(pdf.pages)} 页")
                                for page_idx, page in enumerate(pdf.pages, 1):
                                    try:
                                        page_text = page.extract_text()
                                        if page_text is None:
                                            logger.debug(f"pdfplumber: PDF第{page_idx}页 extract_text() 返回 None")
                                            continue
                                        
                                        page_text = page_text.strip()
                                        if not page_text:
                                            logger.debug(f"pdfplumber: PDF第{page_idx}页文本为空")
                                            continue
                                        
                                        # 清理文本
                                        page_text = ' '.join(page_text.split())
                                        
                                        # 按段落分割
                                        paragraphs = [p.strip() for p in page_text.split('\n\n') if p.strip()]
                                        if not paragraphs:
                                            paragraphs = [p.strip() for p in page_text.split('\n') if p.strip()]
                                        
                                        for para in paragraphs:
                                            if para and len(para) > 10:
                                                texts.append(para)
                                                metadata_list.append({
                                                    'title': f'{filename} - 第{page_idx}页',
                                                    'source': '文件上传',
                                                    'file_name': filename,
                                                    'file_type': 'pdf',
                                                    'page': page_idx,
                                                })
                                        
                                        if page_idx == 1 and texts:
                                            logger.info(f"pdfplumber: PDF第1页成功提取文本，长度: {len(page_text)}")
                                    except Exception as e:
                                        logger.warning(f"pdfplumber提取PDF第{page_idx}页失败: {e}")
                                        continue
                        except ImportError:
                            logger.error("PDF处理需要安装 pdfplumber: pip install pdfplumber")
                            return [], []
                        except Exception as e:
                            logger.error(f"pdfplumber处理也失败: {e}", exc_info=True)
                            return [], []
                    
                    # 如果PyPDF2没有提取到文本，尝试使用pdfplumber
                    if not texts:
                        logger.info(f"PyPDF2未提取到文本，尝试使用pdfplumber: {filename}")
                        try:
                            import pdfplumber
                            with pdfplumber.open(file_path) as pdf:
                                logger.info(f"使用pdfplumber处理PDF，共有 {len(pdf.pages)} 页")
                                for page_idx, page in enumerate(pdf.pages, 1):
                                    try:
                                        page_text = page.extract_text()
                                        if page_text is None:
                                            continue
                                        
                                        page_text = page_text.strip()
                                        if not page_text:
                                            continue
                                        
                                        # 清理文本
                                        page_text = ' '.join(page_text.split())
                                        
                                        # 按段落分割
                                        paragraphs = [p.strip() for p in page_text.split('\n\n') if p.strip()]
                                        if not paragraphs:
                                            paragraphs = [p.strip() for p in page_text.split('\n') if p.strip()]
                                        
                                        for para in paragraphs:
                                            if para and len(para) > 10:
                                                texts.append(para)
                                                metadata_list.append({
                                                    'title': f'{filename} - 第{page_idx}页',
                                                    'source': '文件上传',
                                                    'file_name': filename,
                                                    'file_type': 'pdf',
                                                    'page': page_idx,
                                                })
                                        
                                        if page_idx == 1 and texts:
                                            logger.info(f"pdfplumber: PDF第1页成功提取文本，长度: {len(page_text)}")
                                    except Exception as e:
                                        logger.warning(f"pdfplumber提取PDF第{page_idx}页失败: {e}")
                                        continue
                        except ImportError:
                            logger.warning("pdfplumber未安装，无法尝试备用方法")
                        except Exception as e:
                            logger.warning(f"pdfplumber处理失败: {e}")
                    
                    if not texts:
                        logger.warning(f"PDF文件没有提取到文本: {filename}，可能是扫描版PDF或格式特殊")
                        return [], []
                    else:
                        logger.info(f"PDF文件 {filename} 成功提取 {len(texts)} 条文本")
                        
                except Exception as e:
                    logger.error(f"处理PDF文件失败 {filename}: {e}", exc_info=True)
                    return [], []
            
            else:
                # 其他文件类型，尝试作为文本读取
                logger.info(f"尝试作为文本文件处理: {filename}, 扩展名: {suffix or '无'}")
                try:
                    # 尝试多种编码
                    encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1']
                    content = None
                    for encoding in encodings:
                        try:
                            with open(file_path, 'r', encoding=encoding) as f:
                                content = f.read()
                            logger.info(f"成功使用编码 {encoding} 读取文件")
                            break
                        except UnicodeDecodeError:
                            continue
                    
                    if content is None:
                        logger.warning(f"无法使用任何编码读取文件: {filename}")
                        return [], []
                    
                    # 如果内容为空，返回空列表
                    if not content.strip():
                        logger.warning(f"文件内容为空: {filename}")
                        return [], []
                    
                    # 按段落分割
                    paragraphs = [p.strip() for p in content.split('\n\n') if p.strip()]
                    # 如果没有段落，按行分割
                    if not paragraphs:
                        paragraphs = [p.strip() for p in content.split('\n') if p.strip()]
                    
                    for para in paragraphs:
                        if para:  # 确保段落不为空
                            texts.append(para)
                            metadata_list.append({
                                'title': filename,
                                'source': '文件上传',
                                'file_name': filename,
                                'file_type': suffix[1:] if suffix else 'unknown',
                            })
                    
                    if not texts:
                        logger.warning(f"文件处理后没有提取到文本: {filename}")
                        return [], []
                        
                except Exception as e:
                    logger.warning(f"无法处理文件类型 {suffix}: {e}", exc_info=True)
                    return [], []
            
            logger.info(f"成功处理文件 {filename}，提取 {len(texts)} 条文本")
            return texts, metadata_list
            
        except Exception as e:
            logger.error(f"处理文件失败 {filename}: {e}")
            return [], []
    
    def save_file(self, file, filename):
        """保存上传的文件"""
        file_path = self.upload_dir / filename
        # 如果文件已存在，添加时间戳
        if file_path.exists():
            from datetime import datetime
            name_part = file_path.stem
            ext_part = file_path.suffix
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{name_part}_{timestamp}{ext_part}"
            file_path = self.upload_dir / filename
        
        file.save(str(file_path))
        return file_path

